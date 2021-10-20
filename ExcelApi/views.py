import re
from django.shortcuts import render
from openpyxl.workbook.workbook import Workbook
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .forms import Userform
from django.views import View

from .models import UserInfo
from .Serializer import InfoSerializer

from openpyxl import load_workbook
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render

import sqlite3
db = sqlite3
# Create your views here.
# This is particularly cute view allows a user to post and get json of information in  our
# neat little database :)

class indexView(View):

    def get(self, request):
        return render(request, 'pages/index.html', {'form': Userform})

def req_resolution(request):
    # if this is a POST request we need to process the form data
    if request.method == 'POST':
        # create a form instance and populate it with data from the request:
        form = Userform(request.POST)
        # check whether it's valid:
        if form.is_valid():
            # process the data in form.cleaned_data as required
            # ...
            # redirect to a new URL:
            return HttpResponseRedirect('pages/index.html')

    # if a GET (or any other method) we'll create a blank form
    else:
        form = Userform()

    return render(request, 'pages/index.html', {'form': Userform})

def pop_excel(data):
    try:
        wb = load_workbook('User_Info.xlsx')
        sheet = wb.active
        User_db = db.connect('UserInfo.sqlite3')

        sheet['A1'] = 'Name & Surname'
        sheet['B1'] = 'Contact Number'
        sheet['C1'] = 'Age'
        sheet['D1'] = 'Gender'
        sheet['E1'] = 'Race'
        sheet['F1'] = 'Area'
        sheet['G1'] = 'Gender'
        sheet['H1'] = 'Email'

        sheet['A'+ str(data['id'] + 1)] = data['Name'] + ' ' + data['Surname']
        sheet['B'+ str(data['id'] + 1)] = data['Contact_Number']
        sheet['C'+ str(data['id'] + 1)] = data['Age']
        sheet['D'+ str(data['id'] + 1)] = data['Gender']
        sheet['E'+ str(data['id'] + 1)] = data['Race']
        sheet['F'+ str(data['id'] + 1)] = data['Area']
        sheet['G'+ str(data['id'] + 1)] = data['Gender']
        sheet['H'+ str(data['id'] + 1)] = data['Email']
        
        wb.save('UserInfo.xlsx')
    except FileNotFoundError:
        wb = Workbook()
        wb.save('UserInfo.xlsx')
        sheet = wb.active
        User_db = db.connect('UserInfo.sqlite3')

        sheet['A1'] = 'Name & Surname'
        sheet['B1'] = 'Contact Number'
        sheet['C1'] = 'Age'
        sheet['D1'] = 'Gender'
        sheet['E1'] = 'Race'
        sheet['F1'] = 'Area'
        sheet['G1'] = 'Gender'
        sheet['H1'] = 'Email'

        sheet['A'+ str(data['id'] + 1)] = data['Name'] + ' ' + data['Surname']
        sheet['B'+ str(data['id'] + 1)] = data['Contact_Number']
        sheet['C'+ str(data['id'] + 1)] = data['Age']
        sheet['D'+ str(data['id'] + 1)] = data['Gender']
        sheet['E'+ str(data['id'] + 1)] = data['Race']
        sheet['F'+ str(data['id'] + 1)] = data['Area']
        sheet['G'+ str(data['id'] + 1)] = data['Gender']
        sheet['H'+ str(data['id'] + 1)] = data['Email']
        
        wb.save('UserInfo.xlsx')

class UserInfoView(APIView):

    def get(self, request, format=None):
        Info = UserInfo.objects.all()
        serializer = InfoSerializer(Info, many=True)
        return  Response(serializer.data)

    def post(self, request):
        form = Userform(request.POST)
        print(request.data)
        if form.is_valid():
            pop_excel(request.data)
            new_post = UserInfo(
                U_Name = request.data['Name'],
                U_Surname = request.data['Surname'],
                U_Age = request.data['Age'],
                U_ContactNo = request.data['Contact_Number'],
                U_Gender = request.data['Gender'],
                U_Race = request.data['Race'],
                U_Area = request.data['Area'],
                U_Email = request.data['Email'],
            )
            new_post.save()
            Response(request.data, status=status.HTTP_201_CREATED)
            return HttpResponseRedirect('..', {'form': Userform})

        serializer = InfoSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
